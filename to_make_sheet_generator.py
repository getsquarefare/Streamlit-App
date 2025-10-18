from pyairtable.formulas import match
from pyairtable import Table
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from dotenv import load_dotenv
import streamlit as st
from io import BytesIO
import logging
import ast
import json
from openai import OpenAI

client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])
import os
from exceptions import AirTableError
from store_access import new_database_access
# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
VIEW = "viw4WN1XsjMnHwMkt"

def parse_sauce_multiplier(sauce_field):
    """Parse sauce multiplier from sauce field (e.g., 'Tomato Sauce (2 x sauce)' -> 2)"""
    import re

    # Look for pattern like "(2 x sauce)", "(3 x sauce)", etc.
    multiplier_match = re.search(r'\((\d+)\s*x\s*sauce\)', sauce_field, re.IGNORECASE)
    if multiplier_match:
        return int(multiplier_match.group(1))

    # If no multiplier found, default to 1
    return 1

def parse_modified_recipe_details(modified_recipe_text):
    """Parse the Modified Recipe Details field which contains ingredient names and gram amounts"""
    if not modified_recipe_text:
        return {}

    try:
        # Try to parse as JSON first
        if isinstance(modified_recipe_text, str):
            # Handle both JSON and Python dict string formats
            if modified_recipe_text.startswith('{'):
                try:
                    return json.loads(modified_recipe_text)
                except json.JSONDecodeError:
                    # Try parsing as Python literal
                    return ast.literal_eval(modified_recipe_text)
            else:
                return {}
        elif isinstance(modified_recipe_text, dict):
            return modified_recipe_text
        else:
            return {}
    except (ValueError, SyntaxError) as e:
        logger.warning(f"Could not parse Modified Recipe Details: {modified_recipe_text}, Error: {e}")
        return {}

def identify_fruits_with_openai(ingredient_names):
    """Use OpenAI to identify which ingredients are fruits"""
    try:
        # Get OpenAI API key from environment or streamlit secrets
        openai_api_key = st.secrets.get("OPENAI_API_KEY") if hasattr(st, 'secrets') else None
        if not openai_api_key:
            openai_api_key = os.getenv("OPENAI_API_KEY")

        if not openai_api_key:
            logger.warning("OpenAI API key not found, skipping fruit identification")
            return set()


        # Create prompt for fruit identification
        prompt = f"""
        Given the following list of food ingredients, identify which ones are fruits.
        Return only the fruit names as a comma-separated list, nothing else. Dont change any ingredient names.
        
        Ingredients: {', '.join(ingredient_names)}
        
        Fruits:"""

        response = client.responses.create(model="gpt-5-nano",
        input = prompt,
        reasoning={ 
            "effort": "minimal"},
        text={
            "verbosity": "low"
        })

        fruit_response = response.output_text.strip()
        fruits = {fruit.strip() for fruit in fruit_response.split(',') if fruit.strip()}

        logger.info(f"Identified fruits: {fruits}")
        return fruits

    except Exception as e:
        logger.warning(f"Error identifying fruits with OpenAI: {str(e)}")
        return set()


def group_ingredients_by_component(db,client_servings):
    """Group all ingredients by component type from client servings using Modified Recipe Details"""
    ingredient_summary = {
        'Starch': {},
        'Breakfast Meat': {},  # Separate section for breakfast meats
        'Meat': {},  # Lunch/dinner meats
        'Sauce': {},
        'Garnish': {},  # Will store garnish combinations and their meal counts
        'Veggie': {},
        'Snack': {}  # New section for snacks
    }

    # Special handling for garnish combinations
    garnish_combinations = {}  # Will store unique combinations and their meal counts

    # Special handling for snacks
    snack_dishes = {}  # Will store snack dishes and their total grams
    snack_ingredients = set()  # Track all ingredients that come from snacks


    for serving in client_servings:
        fields = serving['fields']

        # Get meal type from MEAL field
        meal_field = fields.get('Meal Portion (from Linked OrderItem)', '')
         # Get the dish name
        dish_name = fields.get('Dish', ['Unknown'])[0] if fields.get('Dish') else 'Unknown'


        if isinstance(meal_field, list):
            meal_type = meal_field[0] if meal_field else ''
        else:
            meal_type = meal_field if meal_field else ''

        if 'pre-set' in meal_type.lower() or 'add on' in meal_type.lower():
            meal_type = fields.get('Meal Portion (from Linked OrderItem)', '')[0]

        is_snack = 'snack' in meal_type.lower() 
        is_yogurt_breakfast = ('yogurt' in dish_name.lower() or 'parfait' in dish_name.lower()) and 'breakfast' in meal_type.lower()

        # Parse Modified Recipe Details to get actual ingredients and their gram amounts
        modified_recipe = parse_modified_recipe_details(fields.get('Modified Recipe Details', ''))

        if not modified_recipe:
            continue

        # Get ingredient names from each component column to determine component mapping
        starch_names = fields.get('Starch', '').replace('\n', '').split(', ')
        meat_names = fields.get('Meat', '').replace('\n', '').split(', ')
        sauce_field = fields.get('Sauce', '')
        sauce_names = sauce_field.replace('\n', '').replace(' (2 x sauce)', '').replace(' (3 x sauce)', '').replace(' (4 x sauce)', '').split(', ')
        garnish_names = fields.get('Garnish', '').replace('\n', '').split(', ')
        veggie_names = fields.get('Veggies', '').replace('\n', '').split(', ')

        # Parse sauce multiplier from the original sauce field
        sauce_multiplier = parse_sauce_multiplier(sauce_field)

        # Create a mapping of ingredient name to component type
        component_mapping = {}
        for name in starch_names:
            if name != '':
                component_mapping[name.strip()] = 'Starch'
        for name in meat_names:
            if name != '':
                component_mapping[name.strip()] = 'Meat'
        for name in sauce_names:
            if name != '':
                component_mapping[name.strip()] = 'Sauce'
        for name in garnish_names:
            if name != '':
                component_mapping[name.strip()] = 'Garnish'
        for name in veggie_names:
            if name != '':
                component_mapping[name.strip()] = 'Veggie'

       
        # Process each ingredient in Modified Recipe Details
        total_grams = 0
        for ingredient_name, grams in modified_recipe.items():

            # Remove ID prefix from ingredient name (e.g. "16137 Hummus" -> "Hummus")
            clean_ingredient_name = ' '.join(ingredient_name.split(' ')[1:]).strip()

            # Track ingredients that come from snacks
            if is_snack:
                if dish_name not in snack_dishes:
                    snack_dishes[dish_name] = 0
                snack_dishes[dish_name] += grams
                continue

            # Determine component type based on mapping
            component_type = component_mapping.get(clean_ingredient_name)

            if not component_type:
                logger.warning(f"Could not determine component type for ingredient: '{clean_ingredient_name}'")
                logger.info(f"Available component mappings: {list(component_mapping.keys())}")
                continue

            if component_type == 'Garnish':
                # For garnish, we'll handle it separately
                continue

            grams_float = float(grams)

            # Apply raw/cooked conversion factor
            conversion_factor = db.get_ingredient_conversion_factor(ingredient_name.strip())
            final_grams = grams_float / conversion_factor if conversion_factor != 0 else grams_float

            total_grams += final_grams

            # Special handling for meat - exclude snack meats and group by meal type
            if component_type == 'Meat':

                # Determine if this is breakfast meat
                is_breakfast = 'breakfast' in meal_type.lower()
                target_section = 'Breakfast Meat' if is_breakfast else 'Meat'

                if clean_ingredient_name not in ingredient_summary[target_section]:
                    ingredient_summary[target_section][clean_ingredient_name] = {
                        'total_grams': 0
                    }

                ingredient_summary[target_section][clean_ingredient_name]['total_grams'] += final_grams
            
            # Special handling for sauce - count quantities instead of grams
            elif component_type == 'Sauce':
                if clean_ingredient_name not in ingredient_summary[component_type]:
                    ingredient_summary[component_type][clean_ingredient_name] = {
                        'total_count': 0
                    }

                # Add the sauce multiplier count (from "(2 x sauce)" notation)
                ingredient_summary[component_type][clean_ingredient_name]['total_count'] += sauce_multiplier

            # Skip starch from yogurt breakfast, add it to snacks
            elif component_type == 'Starch' and is_yogurt_breakfast:
                if clean_ingredient_name not in snack_dishes:
                    snack_dishes[clean_ingredient_name] = 0
                snack_dishes[clean_ingredient_name] += final_grams
                continue

            else:
                if clean_ingredient_name not in ingredient_summary[component_type]:
                    ingredient_summary[component_type][clean_ingredient_name] = {
                        'total_grams': 0
                    }

                ingredient_summary[component_type][clean_ingredient_name]['total_grams'] += final_grams

        # Handle garnish combinations
        if garnish_names and any(name.strip() for name in garnish_names):
            if dish_name not in garnish_combinations:
                garnish_combinations[dish_name] = {
                    'garnishes': set(),
                    'meal_count': 0
                }
            garnish_combinations[dish_name]['garnishes'].update(name.strip() for name in garnish_names if name.strip())
            garnish_combinations[dish_name]['meal_count'] += 1

    # Add garnish combinations to the summary
    ingredient_summary['Garnish'] = {
        dish: {
            'garnish_combo': ', '.join(sorted(garnish_data['garnishes'])),
            'meal_count': garnish_data['meal_count']
        }
        for dish, garnish_data in garnish_combinations.items()
    }

    # Add snack dishes to the summary
    ingredient_summary['Snack'] = {
        dish: {'total_grams': grams}
        for dish, grams in snack_dishes.items()
    }

    # Consolidate duplicate ingredients between Breakfast Meat and Meat sections
    breakfast_meats = ingredient_summary['Breakfast Meat']
    lunch_dinner_meats = ingredient_summary['Meat']

    # Find duplicated meats and consolidate them into Breakfast Meat
    duplicates_to_remove = []
    for ingredient_name in lunch_dinner_meats:
        if ingredient_name in breakfast_meats:
            # Add the lunch/dinner meat quantity to breakfast meat
            breakfast_meats[ingredient_name]['total_grams'] += lunch_dinner_meats[ingredient_name]['total_grams']
            # Mark for removal from lunch/dinner meat
            duplicates_to_remove.append(ingredient_name)
            logger.info(f"Consolidated duplicate ingredient '{ingredient_name}' into Breakfast Meat section")

    # Remove duplicates from lunch/dinner meat section
    for ingredient_name in duplicates_to_remove:
        del lunch_dinner_meats[ingredient_name]

    # Debug: Log final summary
    for component, ingredients in ingredient_summary.items():
        if ingredients:
            logger.info(f"{component} section has {len(ingredients)} items: {list(ingredients.keys())}")
        else:
            logger.info(f"{component} section is empty")
    return ingredient_summary

def cluster_veggies_by_preparation(veggies_dict):
    """Cluster vegetables by preparation method based on ingredient names"""
    clusters = {
        'Roasted / Charred / Sauteed': [],
        'Steamed': [],
        'Raw': [],
        'Unseasoned': [],
        'Other': []
    }

    for veggie_name, veggie_data in veggies_dict.items():
        name_lower = veggie_name.lower()

        # Check for roasted/charred/sauteed
        if any(keyword in name_lower for keyword in ['roasted', 'charred', 'sauteed', 'sautéed']):
            clusters['Roasted / Charred / Sauteed'].append((veggie_name, veggie_data))
        # Check for steamed
        elif 'steamed' in name_lower:
            clusters['Steamed'].append((veggie_name, veggie_data))
        # Check for raw
        elif 'raw' in name_lower:
            clusters['Raw'].append((veggie_name, veggie_data))
        # Check for unseasoned
        elif 'unseasoned' in name_lower:
            clusters['Unseasoned'].append((veggie_name, veggie_data))
        # Everything else goes to Other
        else:
            clusters['Other'].append((veggie_name, veggie_data))

    # Sort each cluster by quantity (grams) in descending order
    for cluster_name in clusters:
        clusters[cluster_name].sort(key=lambda x: x[1]['total_grams'], reverse=True)

    return clusters

def sort_breakfast_meats_with_openai(breakfast_meat_list):
    """Use OpenAI to intelligently sort breakfast meats, prioritizing eggs, oats, and yogurt"""
    if not breakfast_meat_list:
        return breakfast_meat_list

    try:
        # Extract just the names for OpenAI
        meat_names = [item[0] for item in breakfast_meat_list]

        # Create prompt for breakfast protein sorting
        prompt = f"""
        Given the following list of breakfast protein/food items, sort them by grouping similar types together in this priority order:
        1. Egg-related proteins first (eggs, egg whites, scrambled eggs, etc.)
        2. Oats/oatmeal items second
        3. Yogurt items third
        4. Other items grouped by type (cluster similar items together)

        Within each group, maintain alphabetical order.
        Return ONLY the sorted list of names, one per line, in the exact same format as provided. Do not add numbering, bullets, or any extra text.

        Items to sort:
        {chr(10).join(meat_names)}

        Sorted items:"""

        response = client.responses.create(
            model="gpt-5-nano",
            input=prompt,
            reasoning={"effort": "minimal"},
            text={"verbosity": "low"}
        )

        sorted_names = [name.strip() for name in response.output_text.strip().split('\n') if name.strip()]

        # Create a mapping of name to original tuple
        name_to_tuple = {item[0]: item for item in breakfast_meat_list}

        # Rebuild the sorted list
        sorted_list = []
        for name in sorted_names:
            if name in name_to_tuple:
                sorted_list.append(name_to_tuple[name])

        # Add any items that weren't in the sorted response (fallback)
        sorted_names_set = set(sorted_names)
        for item in breakfast_meat_list:
            if item[0] not in sorted_names_set:
                sorted_list.append(item)
                logger.warning(f"OpenAI did not include '{item[0]}' in breakfast sort response, appending to end")

        logger.info(f"Sorted {len(breakfast_meat_list)} breakfast items by type using OpenAI")
        return sorted_list

    except Exception as e:
        logger.warning(f"Error sorting breakfast meats with OpenAI: {str(e)}, falling back to alphabetical sort")
        # Fallback to alphabetical sorting
        return sorted(breakfast_meat_list, key=lambda x: x[0])

def sort_meats_by_protein_type_with_openai(meat_list):
    """Use OpenAI to intelligently sort meats by protein type (chicken, salmon, tofu, etc.)"""
    if not meat_list:
        return meat_list

    try:
        # Extract just the names for OpenAI
        meat_names = [item[0] for item in meat_list]

        # Create prompt for protein type sorting
        prompt = f"""
        Given the following list of meat/protein items, sort them by protein type in this specific order:
        1. Chicken items first
        2. Salmon items second
        3. Tofu items third
        4. Other meats (beef, steak, pork, lamb, etc.)
        5. Other fish/seafood (shrimp, tuna, cod, etc.)
        6. Other vegan/vegetarian proteins (beans, tempeh, seitan, etc.)

        Return ONLY the sorted list of names, one per line, in the exact same format as provided. Do not add numbering, bullets, or any extra text.

        Items to sort:
        {';'.join(meat_names)}

        Sorted items:"""

        response = client.responses.create(
            model="gpt-5-nano",
            input=prompt,
            reasoning={"effort": "minimal"},
            text={"verbosity": "low"}
        )

        sorted_names = [name.strip() for name in response.output_text.strip().split('\n') if name.strip()]

        # Create a mapping of name to original tuple
        name_to_tuple = {item[0]: item for item in meat_list}

        # Rebuild the sorted list
        sorted_list = []
        for name in sorted_names:
            if name in name_to_tuple:
                sorted_list.append(name_to_tuple[name])

        # Add any items that weren't in the sorted response (fallback)
        sorted_names_set = set(sorted_names)
        for item in meat_list:
            if item[0] not in sorted_names_set:
                sorted_list.append(item)
                logger.warning(f"OpenAI did not include '{item[0]}' in sort response, appending to end")

        logger.info(f"Sorted {len(meat_list)} meat items by protein type using OpenAI")
        return sorted_list

    except Exception as e:
        logger.warning(f"Error sorting meats with OpenAI: {str(e)}, falling back to alphabetical sort")
        # Fallback to alphabetical sorting
        return sorted(meat_list, key=lambda x: x[0])

def cluster_meats_by_type(meats_dict):
    """Cluster lunch/dinner meats by type and sourcing based on ingredient names"""
    clusters = {
        'Flavored Proteins': [],
        'Flavored Proteins - Wild': [],
        'Grilled Proteins': [],
        'Grilled Proteins - Wild': [],
        'Unseasoned Proteins': []
    }

    for meat_name, meat_data in meats_dict.items():
        name_lower = meat_name.lower()

        # Check if it's wild/organic/grassfed
        is_wild = any(keyword in name_lower for keyword in ['organic', 'wild', 'grassfed', 'grass-fed', 'grass fed'])

        # Check for unseasoned
        if 'unseasoned' in name_lower:
            clusters['Unseasoned Proteins'].append((meat_name, meat_data))
        # Check for grilled
        elif 'grilled' in name_lower:
            if is_wild:
                clusters['Grilled Proteins - Wild'].append((meat_name, meat_data))
            else:
                clusters['Grilled Proteins'].append((meat_name, meat_data))
        # Everything else is flavored
        else:
            if is_wild:
                clusters['Flavored Proteins - Wild'].append((meat_name, meat_data))
            else:
                clusters['Flavored Proteins'].append((meat_name, meat_data))

    # Sort each cluster using OpenAI to intelligently group by protein type
    for cluster_name in clusters:
        clusters[cluster_name] = sort_meats_by_protein_type_with_openai(clusters[cluster_name])

    return clusters

def create_to_make_sheet_excel(ingredient_summary):
    """Create Excel file with to-make sheet formatted for printing"""
    try:
        # current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = BytesIO()

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "To-Make Sheet"

        # Set up print formatting
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        # Define styles
        header_font = Font(bold=True, size=14)
        component_header_font = Font(bold=True, size=12, color="FFFFFF")
        ingredient_font = Font(size=11)
        detail_font = Font(size=9)

        # Define fills
        component_fills = {
            'Starch': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'Breakfast Meat': PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid"),  # Orange for breakfast meat
            'Meat': PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),  
            'Sauce': PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
            'Garnish': PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
            'Veggie': PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid"),
            'Snack': PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")  # Using same purple as Veggie
        }

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add main header
        """ ws.merge_cells('A1:K1')
        header_cell = ws['A1']
        header_cell.value = f"TO-MAKE SHEET - Created on{datetime.now().strftime('%B %d, %Y')}"
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = thin_border """

        # Initialize row counters for left and right columns
        left_row = 1
        right_row = 1

        # Define section order and their positions
        left_sections = ['Veggie', 'Garnish']
        right_sections = ['Starch', 'Meat']
        bottom_sections = ['Sauce', 'Snack']  # New sections for bottom row

        # Process left column sections
        for component_type in left_sections:
            if component_type not in ingredient_summary or not ingredient_summary[component_type]:
                continue

            ingredients = ingredient_summary[component_type]

            # Column headers for this component
            headers = ['Ingredient', 'Preparation', 'Uncooked (g)','uncooked (lbs)'] if component_type != 'Garnish' else ['Garnishes', '# of Meals']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=left_row, column=col)
                if col == 1:  # First column (Ingredient)
                    cell.value = f"{component_type.capitalize()}"
                else:
                    cell.value = header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.border = thin_border
                cell.fill = component_fills[component_type]
            left_row += 1

            if component_type == 'Garnish':
                # Sort garnish combinations by meal count (descending)
                sorted_items = sorted(ingredients.items(),
                                    key=lambda x: (x[0].split()[-1], -x[1]['meal_count']),
                                    reverse=True)

                for item_name, data in sorted_items:
                    # Garnish row
                    ws.cell(row=left_row, column=1, value=data['garnish_combo']).font = ingredient_font
                    ws.cell(row=left_row, column=2, value=data['meal_count']).font = ingredient_font

                    # Apply borders to all cells in the row
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=left_row, column=col).border = thin_border
                        ws.cell(row=left_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

                    left_row += 1
                
            elif component_type == 'Veggie':
                # Special handling for Veggie section - cluster by preparation method
                veggie_clusters = cluster_veggies_by_preparation(ingredients)

                for cluster_name, cluster_items in veggie_clusters.items():
                    # Skip empty clusters
                    if not cluster_items:
                        continue

                    # Cluster sub-header with lighter background
                    ws.cell(row=left_row, column=1, value=cluster_name).font = Font(bold=True, size=12)
                    ws.cell(row=left_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                    ws.merge_cells(start_row=left_row, start_column=1, end_row=left_row, end_column=4)
                    ws.cell(row=left_row, column=1).border = thin_border
                    ws.cell(row=left_row, column=4).border = thin_border
                    ws.cell(row=left_row, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light purple
                    left_row += 1

                    # Add vegetables in this cluster
                    for item_name, data in cluster_items:
                        ws.cell(row=left_row, column=1, value=item_name).font = ingredient_font
                        ws.cell(row=left_row, column=3, value=round(data['total_grams'], 1)).font = ingredient_font
                        ws.cell(row=left_row, column=4, value=round(data['total_grams']/453.592, 2)).font = ingredient_font

                        # Apply borders to all cells in the row
                        for col in range(1, len(headers) + 1):
                            ws.cell(row=left_row, column=col).border = thin_border
                            ws.cell(row=left_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                        left_row += 1
            else:
                # Sort ingredients: first by name (for grouping), then by total grams (descending)
                sorted_items = sorted(ingredients.items(),
                                    key=lambda x: (x[0].split()[-1], -x[1]['total_grams']))

                for item_name, data in sorted_items:
                    # Ingredient row
                    ws.cell(row=left_row, column=1, value=item_name).font = ingredient_font
                    ws.cell(row=left_row, column=3, value=round(data['total_grams'], 1)).font = ingredient_font
                    ws.cell(row=left_row, column=4, value=round(data['total_grams']/453.592, 2)).font = ingredient_font

                    # Apply borders to all cells in the row
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=left_row, column=col).border = thin_border
                        ws.cell(row=left_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

                    left_row += 1
            left_row += 1

        right_section_starting_column = 6
        # Process right column sections
        for component_type in right_sections:
            if component_type == 'Meat':
                # Special handling for Meat section with sub-sections
                breakfast_meats = ingredient_summary.get('Breakfast Meat', {})
                lunch_dinner_meats = ingredient_summary.get('Meat', {})

                if not breakfast_meats and not lunch_dinner_meats:
                    continue

                # Main Meat header
                headers = ['Ingredient', 'Preparation', 'Uncooked (g)','uncooked (lbs)']

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=right_row, column=right_section_starting_column + col - 1)
                    if col == 1:  # First column (Ingredient)
                        cell.value = "Meat"
                    else:
                        cell.value = header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.border = thin_border
                    cell.fill = component_fills['Meat']
                right_row += 1

                # Breakfast Meat sub-section
                if breakfast_meats:
                    # Breakfast sub-header
                    ws.cell(row=right_row, column=right_section_starting_column, value="Breakfast:").font = Font(bold=True, size=12)
                    for col in range(right_section_starting_column, right_section_starting_column + len(headers)):
                        ws.cell(row=right_row, column=col).border = thin_border
                        ws.cell(row=right_row, column=col).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    right_row += 1

                    # Sort breakfast meats using OpenAI to intelligently group by type (eggs, oats, yogurt, etc.)
                    breakfast_items_list = [(name, data) for name, data in breakfast_meats.items()]
                    sorted_breakfast = sort_breakfast_meats_with_openai(breakfast_items_list)

                    for item_name, data in sorted_breakfast:
                        ws.cell(row=right_row, column=right_section_starting_column, value=item_name).font = ingredient_font
                        ws.cell(row=right_row, column=right_section_starting_column + 2, value=round(data['total_grams'], 1)).font = ingredient_font
                        ws.cell(row=right_row, column=right_section_starting_column + 3, value=round(data['total_grams']/453.592, 2)).font = ingredient_font

                        for col in range(right_section_starting_column, right_section_starting_column + len(headers)):
                            ws.cell(row=right_row, column=col).border = thin_border
                            ws.cell(row=right_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

                        right_row += 1

                # Lunch/Dinner Meat sub-section
                if lunch_dinner_meats:
                    # Lunch/Dinner sub-header
                    ws.cell(row=right_row, column=right_section_starting_column, value="Lunch/Dinner:").font = Font(bold=True, size=12)
                    for col in range(right_section_starting_column, right_section_starting_column + len(headers)):
                        ws.cell(row=right_row, column=col).border = thin_border
                        ws.cell(row=right_row, column=col).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    right_row += 1

                    # Cluster lunch/dinner meats by type
                    meat_clusters = cluster_meats_by_type(lunch_dinner_meats)

                    for cluster_name, cluster_items in meat_clusters.items():
                        # Skip empty clusters
                        if not cluster_items:
                            continue

                        # Cluster sub-header with lighter background
                        ws.cell(row=right_row, column=right_section_starting_column, value=cluster_name).font = Font(bold=True, size=12)
                        ws.merge_cells(start_row=right_row, start_column=right_section_starting_column, end_row=right_row, end_column=right_section_starting_column + 3)
                        ws.cell(row=right_row, column=right_section_starting_column).alignment = Alignment(horizontal='left', vertical='center')
                        ws.cell(row=right_row, column=right_section_starting_column+3).border = thin_border
                        ws.cell(row=right_row, column=right_section_starting_column).border = thin_border
                        ws.cell(row=right_row, column=right_section_starting_column).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
                        right_row += 1

                        # Add meats in this cluster
                        for item_name, data in cluster_items:
                            ws.cell(row=right_row, column=right_section_starting_column, value=item_name).font = ingredient_font
                            ws.cell(row=right_row, column=right_section_starting_column + 2, value=round(data['total_grams'], 1)).font = ingredient_font
                            ws.cell(row=right_row, column=right_section_starting_column + 3, value=round(data['total_grams']/453.592, 2)).font = ingredient_font

                            for col in range(right_section_starting_column, right_section_starting_column + len(headers)):
                                ws.cell(row=right_row, column=col).border = thin_border
                                ws.cell(row=right_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

                            right_row += 1

            else:
                # Handle other sections normally (like Starch)
                if component_type not in ingredient_summary or not ingredient_summary[component_type]:
                    continue

                ingredients = ingredient_summary[component_type]

                # Column headers for this component
                headers = ['Ingredient', 'Preparation', 'Uncooked (g)','uncooked (lbs)']

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=right_row, column=col + right_section_starting_column - 1)  # Start from column G (7)
                    if col == 1:  # First column (Ingredient)
                        cell.value = f"{component_type.capitalize()}"
                    else:
                        cell.value = header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.border = thin_border
                    cell.fill = component_fills[component_type]
                right_row += 1

                # Sort ingredients by total grams (descending)
                sorted_items = sorted(ingredients.items(),
                                    key=lambda x: x[1]['total_grams'], reverse=True)

                for item_name, data in sorted_items:
                    # Ingredient row
                    ws.cell(row=right_row, column=right_section_starting_column, value=item_name).font = ingredient_font
                    ws.cell(row=right_row, column=right_section_starting_column + 2, value=round(data['total_grams'], 1)).font = ingredient_font
                    ws.cell(row=right_row, column=right_section_starting_column + 3, value=round(data['total_grams']/453.592, 2)).font = ingredient_font
                    # Apply borders to all cells in the row
                    for col in range(right_section_starting_column, right_section_starting_column + len(headers)):
                        ws.cell(row=right_row, column=col).border = thin_border
                        ws.cell(row=right_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

                    right_row += 1
                right_row += 1

        # Process bottom sections (Sauce and Snack)
        bottom_row = right_row + 1

        # Process Sauce section
        if 'Sauce' in ingredient_summary and ingredient_summary['Sauce']:
            sauce_headers = ['Sauce to Make', '# of Servings']
            start_col = right_section_starting_column  # Start from column G

            for col, header in enumerate(sauce_headers, 1):
                cell = ws.cell(row=bottom_row, column=col + start_col - 1)
                if col == 1:  # First column
                    cell.value = "Sauce"
                else:
                    cell.value = header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.border = thin_border
                cell.fill = component_fills['Sauce']

            # Sort sauces: first by name (for grouping), then by total count (descending)
            sorted_items = sorted(ingredient_summary['Sauce'].items(), 
                                key=lambda x: (x[0].split()[-1], -x[1]['total_count']))

            for item_name, data in sorted_items:
                # Row
                ws.cell(row=bottom_row + 1, column=start_col, value=item_name).font = ingredient_font
                ws.cell(row=bottom_row + 1, column=start_col + 1, value=data['total_count']).font = ingredient_font

                # Apply borders to all cells in the row
                for col in range(start_col, start_col + len(sauce_headers)):
                    ws.cell(row=bottom_row + 1, column=col).border = thin_border
                    ws.cell(row=bottom_row + 1, column=col).alignment = Alignment(wrap_text=True, vertical='top')

                bottom_row += 1

        # Process Snack section (starting at the same row as Sauce)
        if 'Snack' in ingredient_summary and ingredient_summary['Snack']:
            snack_headers = ['Snack', 'Total Grams (g)','Total Grams (lbs)']
            start_col = right_section_starting_column + 3 

            for col, header in enumerate(snack_headers, 1):
                cell = ws.cell(row=bottom_row - len(ingredient_summary['Sauce'].items()) if 'Sauce' in ingredient_summary else bottom_row, 
                                column=col + start_col - 1)
                if col == 1:  # First column
                    cell.value = "Snack"
                else:
                    cell.value = header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.border = thin_border
                cell.fill = component_fills['Snack']

            # Sort snacks: first by name (for grouping), then by total grams (descending)
            sorted_items = sorted(ingredient_summary['Snack'].items(), 
                                key=lambda x: (x[0].split()[-1], -x[1]['total_grams']))

            for item_name, data in sorted_items:
                # Row
                ws.cell(row=bottom_row - len(ingredient_summary['Sauce'].items()) + 1 if 'Sauce' in ingredient_summary else bottom_row + 1, 
                        column=start_col, value=item_name).font = ingredient_font
                ws.cell(row=bottom_row - len(ingredient_summary['Sauce'].items()) + 1 if 'Sauce' in ingredient_summary else bottom_row + 1, 
                        column=start_col + 1, value=round(data['total_grams'], 1)).font = ingredient_font
                ws.cell(row=bottom_row - len(ingredient_summary['Sauce'].items()) + 1 if 'Sauce' in ingredient_summary else bottom_row + 1, 
                        column=start_col + 2, value=round(data['total_grams']/453.592, 2)).font = ingredient_font
                # Apply borders to all cells in the row
                for col in range(start_col, start_col + len(snack_headers)):
                    ws.cell(row=bottom_row - len(ingredient_summary['Sauce'].items()) + 1 if 'Sauce' in ingredient_summary else bottom_row + 1, 
                            column=col).border = thin_border
                    ws.cell(row=bottom_row - len(ingredient_summary['Sauce'].items()) + 1 if 'Sauce' in ingredient_summary else bottom_row + 1, 
                            column=col).alignment = Alignment(wrap_text=True, vertical='top')

                bottom_row += 1

        # Set row heights
        max_row = max(left_row, right_row, bottom_row)
        for row in range(1, max_row):
            ws.row_dimensions[row].height = 20

        # Save to BytesIO
        wb.save(output)
        output.seek(0)

        logger.info(f"Successfully created to-make sheet Excel file")
        return output

    except Exception as e:
        logger.error(f"Error creating to-make sheet Excel: {str(e)}")
        raise AirTableError(f"Error creating to-make sheet Excel: {str(e)}")

def generate_to_make_sheet(db):
    """Main method to generate the to-make sheet"""
    try:
        logger.info("Starting to-make sheet generation")
        # Get client servings data
        client_servings = db.get_clientservings_data(view=VIEW)
        if not client_servings:
            raise AirTableError("No client servings data found")

        logger.info(f"Found {len(client_servings)} client servings")

        # Group ingredients by component
        ingredient_summary = group_ingredients_by_component(db, client_servings)

        # Log summary
        for component, ingredients in ingredient_summary.items():
            logger.info(f"{component}: {len(ingredients)} unique ingredients")

        # move any eggs from lunch/dinner meat to breakfast meat
        eggs_to_move = []
        for meat in ingredient_summary['Meat']:
            if 'egg' in meat.lower() or 'eggs' in meat.lower():
                eggs_to_move.append(meat)

        for meat in eggs_to_move:
            ingredient_summary['Breakfast Meat'][meat] = ingredient_summary['Meat'][meat]
            del ingredient_summary['Meat'][meat]

        # move fruits from breakfast meat to snack section
        if ingredient_summary['Veggie']:

            veggies_ingredients = list(ingredient_summary['Veggie'].keys())

            # Identify fruits using OpenAI
            fruits = identify_fruits_with_openai(veggies_ingredients)

            # Move fruits from breakfast meat to snack
            fruits_to_move = []
            for ingredient in ingredient_summary['Veggie']:
                if ingredient in fruits:
                    fruits_to_move.append(ingredient)

            for fruit in fruits_to_move:
                ingredient_summary['Snack'][fruit] = ingredient_summary['Veggie'][fruit]
                del ingredient_summary['Veggie'][fruit]

        # Create Excel file
        excel_file = create_to_make_sheet_excel(ingredient_summary)
        return excel_file

    except Exception as e:
        logger.error(f"Error generating to-make sheet: {str(e)}")
        raise AirTableError(f"Error generating to-make sheet: {str(e)}")


if __name__ == "__main__":
    try:
        db = new_database_access()  
        excel_file = generate_to_make_sheet(db)

        # Write the BytesIO content to a file
        filename = f'to_make_sheet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        with open(filename, 'wb') as f:
            f.write(excel_file.getvalue())

        logger.info(f"Successfully generated to-make sheet: {filename}")

    except AirTableError as e:
        logger.critical(f"Application error: {str(e)}")
    except Exception as e:
        logger.critical(f"Unexpected error: {str(e)}")