from openpyxl.styles import PatternFill
import pandas as pd
from datetime import datetime
from io import BytesIO
import re
from exceptions import AirTableError
import logging
from store_access import new_database_access

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def format_output_order_ingredients(db, deleted_ingredients, new_ingredients, final_ingredients_not_in_recommend,tags):
    """Format output for ordered ingredients with deletions and additions"""
    components_output = {
        "Meat": [], 
        "Sauce": [],
        "Starch": [], 
        "Veggies": [], 
        "Garnish": []
    }

    if len(deleted_ingredients) == 0 and len(new_ingredients) == 0:
        return components_output
        
    try:
        # Process deleted ingredients
        for ingredient_id in deleted_ingredients:
            ingredient_details = db.get_ingredient_details_by_rec_id(ingredient_id)
            if ingredient_details:
                ingredient_name = ingredient_details['Ingredient Name']
                ingredient_component = ingredient_details['Component']
                if ingredient_component in components_output:
                    if (ingredient_component == 'Sauce' and 'No Sauce' in tags):
                        components_output[ingredient_component].append("NO SAUCE")
                    elif (ingredient_component == 'Starch' and 'No Starch' in tags):
                        components_output[ingredient_component].append("NO STARCH")
                    elif (ingredient_component == 'Veggies' and 'No Veggies' in tags):
                        components_output[ingredient_component].append("NO VEGGIES")
                    elif (ingredient_component == 'Garnish' and 'No Garnish' in tags):
                        components_output[ingredient_component].append("NO GARNISH")
                    elif (ingredient_component == 'Meat' and 'No Protein' in tags):
                        components_output[ingredient_component].append("NO PROTEIN  ")
                    else:
                        components_output[ingredient_component].append("NO " + ingredient_name.upper())
                    
        # Process new ingredients
        for ingredient_id in new_ingredients:
            ingredient_details = db.get_ingredient_details_by_rec_id(ingredient_id)
            if ingredient_details:
                ingredient_name = ingredient_details['Ingredient Name']
                ingredient_component = ingredient_details['Component']
                if ingredient_component in components_output:
                    if ingredient_id in final_ingredients_not_in_recommend:
                        components_output[ingredient_component].append(ingredient_name.upper()+"(✩)")
                    else:
                        components_output[ingredient_component].append(ingredient_name.upper())
                    
        return components_output
    except Exception as e:
        logger.error(f"Error formatting output for ordered ingredients: {str(e)}")
        raise AirTableError(f"Error formatting output for ordered ingredients: {str(e)}")

def format_output_default_ingredients(db, default_ingredients):
    """Format output for default ingredients"""
    components_output = {
        "Meat": [], 
        "Sauce": [],
        "Starch": [], 
        "Veggies": [], 
        "Garnish": []
    }

    try:
        for ingredient in default_ingredients:
            if not ingredient:
                continue
                
            ingredient_details = db.get_ingredient_details_by_rec_id(ingredient)
            if ingredient_details:
                ingredient_name = ingredient_details['Ingredient Name']
                ingredient_component = ingredient_details['Component']
                if ingredient_component in components_output:
                    components_output[ingredient_component].append(ingredient_name)
                    
        return components_output
    except Exception as e:
        logger.error(f"Error formatting output for default ingredients: {str(e)}")
        raise AirTableError(f"Error formatting output for default ingredients: {str(e)}")

def one_dish_output(db, dish_id):
    """Generate output for one dish including all client servings"""
    try:
        outputs = []
        default_ingredients = db.get_dish_default_ingredients(dish_id)
        dish_product_name = db.get_dish_squarespace_name(dish_id)
        ordered_clientservings = db.get_clientservings_one_dish(dish_id)
        
        default_ingredients_formatted = format_output_default_ingredients(db, default_ingredients)
        default_ingredients_formatted[' '] = dish_id
        default_ingredients_formatted['Position'] = dish_product_name
        
        outputs.append(default_ingredients_formatted)
        
        for client_serving in ordered_clientservings:
            deleted_ingredients = client_serving['fields'].get('All Deletions', [])
            new_ingredients = client_serving['fields'].get('New Ingredients', [])
            ingredients_to_recommend = client_serving['fields'].get('Ingredients To Recommend (from Linked OrderItem)', [])
            final_ingredients_with_user_edits = client_serving['fields'].get('Final Ingredients With User Edits (from Linked OrderItem)', [])
            tags = client_serving['fields'].get('Customization Tags (from To_Match_Client_Nutrition) (from Linked OrderItem)', [])
            
            # Get deleted ingredient names
            deleted_ingredients_names = []
            for ingredient_id in deleted_ingredients:
                ingredient_details = db.get_ingredient_details_by_rec_id(ingredient_id)
                if ingredient_details:
                    deleted_ingredients_names.append(ingredient_details['Ingredient Name'])

            # Get ingredients in final ingredients but not in ingredients to recommend
            final_ingredients_not_in_recommend = []
            for ingredient_id in final_ingredients_with_user_edits:
                if ingredient_id not in ingredients_to_recommend:
                    final_ingredients_not_in_recommend.append(ingredient_id)
            
            # Build output dictionary
            output = {
                ' ': dish_id,
                'Position': client_serving['fields'].get('Position Id', 0),
                'Delivery Date': client_serving['fields']['Delivery Date'],
                'Client': str(client_serving['fields']['Customer Name'][0]),
                'Allergies': client_serving['fields'].get('Nutrition Notes (from Linked OrderItem)', [""])[0],
                'Meal': client_serving['fields'].get('Meal Portion (from Linked OrderItem)', [''])[0].strip(),
                'Sticker': client_serving['fields'].get('Meal Sticker (from Linked OrderItem)', [''])[0],
                #'Dish': client_serving['fields'].get('Dish', [''])[0],
                'All Deletions': deleted_ingredients_names,
                'Portions': 1
            }
            
            # Get components output
            components_output = format_output_order_ingredients(db, deleted_ingredients, new_ingredients, final_ingredients_not_in_recommend,tags)
            
            # Add component amounts
            components_output['Meat'].append(
                '-' if client_serving['fields']['Meat (g)'] == 0 else round(client_serving['fields']['Meat (g)'], 1))
            
            # Handle Sauce with special case for "n x sauce" format
            sauce_value = client_serving['fields']['Sauce (g)']
            sauce_text = client_serving['fields'].get('Sauce', '')
            
            if sauce_value == 0:
                components_output['Sauce'].append('-')
            else:
                sauce_output = str(round(sauce_value, 1))
                # Extract any content in parentheses if it exists
                parentheses_pattern = re.search(r'\((.*?)\)', sauce_text)
                if parentheses_pattern:
                    sauce_output += f" ({parentheses_pattern.group(1)})"
                components_output['Sauce'].append(sauce_output)
            
            components_output['Starch'].append(
                '-' if client_serving['fields']['Starch (g)'] == 0 else round(client_serving['fields']['Starch (g)'], 1))
            components_output['Veggies'].append(
                '-' if client_serving['fields']['Veggies (g)'] == 0 else round(client_serving['fields']['Veggies (g)'], 1))
            components_output['Garnish'].append(
                '-' if client_serving['fields']['Garnish (g)'] == 0 else round(client_serving['fields']['Garnish (g)'], 1))
            
            output.update(components_output)
            outputs.append(output)
            
        return outputs
    except Exception as e:
        logger.error(f"Error generating output for dish ID {dish_id}: {str(e)}")
        raise AirTableError(f"Error generating output for dish ID {dish_id}: {str(e)}")

def generate_formatted_clientservings_onedish(clientservings):
    """Format client servings data for one dish into a dataframe"""
    try:
        def flatten(item):
            if isinstance(item, list):
                return ', '.join(map(str, item))
            return item

        flattened_data = [{k: flatten(v) for k, v in entry.items()} for entry in clientservings]
        df = pd.DataFrame(flattened_data)
        return df
    except Exception as e:
        logger.error(f"Error formatting client servings data: {str(e)}")
        raise AirTableError(f"Error formatting client servings data: {str(e)}")

def consolidated_all_dishes_output(db):
    """Consolidate output for all dishes"""
    try:
        all_clientservings = db.clientserving_table.all(fields=['Dish ID (from Linked OrderItem)'],view='viwgt50kLisz8jx7b')
        if len(all_clientservings) == 0:
            raise AirTableError("No clientservings found in the source view. Please check the view and try again.")
        all_dishes = set()
        all_output = pd.DataFrame()
        
        # Collect all unique dish IDs
        for client_serving in all_clientservings:
            dish_ids = client_serving['fields'].get('Dish ID (from Linked OrderItem)', [])
            if dish_ids:
                all_dishes.add(dish_ids[0])
        all_dishes = sorted(all_dishes)
        # Process each dish
        for dish_id in all_dishes:
            result = one_dish_output(db, dish_id)
            result_df = generate_formatted_clientservings_onedish(result)
            all_output = pd.concat([all_output, result_df], axis=0, ignore_index=True)
            
        return all_output
    except Exception as e:
        logger.error(f"Error consolidating all dishes output: {str(e)}")
        raise AirTableError(f"Error consolidating all dishes output: {str(e)}")

def generate_clientservings_excel(db):
    formatted_clientservings = consolidated_all_dishes_output(db)
    """Generate Excel file with formatted client servings data"""
    try:
        current_date = datetime.now().strftime("%Y%m%d")
        output = BytesIO()
        logger.info('Generating Excel file for formatted client servings')
        # Select and reorder columns
        formatted_clientservings = formatted_clientservings[[
            'Position', 'Client', 'Meal','Portions','Delivery Date', 
            'All Deletions', 'Sauce', 'Garnish', 'Starch', 'Veggies', 
            'Meat', 'Allergies'
        ]]
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            formatted_clientservings.to_excel(writer, index=False, sheet_name='Sheet1')

            # Access the workbook and sheet
            workbook = writer.book
            sheet = workbook['Sheet1']

            # Define fill colors
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # Apply row colors based on NaN index
            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
                if pd.isna(formatted_clientservings.loc[i, 'Delivery Date']):
                    for cell in row:
                        cell.fill = blue_fill

        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        raise AirTableError(f"Error generating Excel file: {str(e)}")


if __name__ == "__main__":
    try:
        db = new_database_access()
        file = generate_clientservings_excel(db)
        
        # Write the BytesIO content to a file
        with open('clientservings_excel_output_'+datetime.now().strftime("%Y%m%d")+'.xlsx', 'wb') as f:
            f.write(file.getvalue())
        logger.info("Successfully generated Excel file")
    except AirTableError as e:
        logger.critical(f"Application error: {str(e)}")
    except Exception as e:
        logger.critical(f"Unexpected error: {str(e)}")